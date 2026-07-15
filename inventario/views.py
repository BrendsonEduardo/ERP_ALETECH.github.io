# inventario/views.py
"""
Views do módulo de Inventário de Ativos de TI.

Todas as views são Function-Based Views (FBVs) protegidas pelo decorator
@login_required, redirecionando para a página de login configurada em settings.
"""

import io
import json
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods

from .forms import AtivoForm, ModeloEquipamentoForm
from .models import Ativo, ModeloEquipamento, Movimentacao, MovimentacaoItem
from comercial.models import Cliente
from django.db import transaction
from django.core.files.base import ContentFile
from .utils import render_to_pdf


# ─────────────────────────────────────────────────────────────────────────────
# ATIVOS
# ─────────────────────────────────────────────────────────────────────────────


@login_required(login_url="login")
def lista_ativos(request):
    """
    Dashboard de ativos com filtros dinâmicos por coluna e KPIs reativos.

    Parâmetros GET aceitos:
      ?q=<texto>        → busca livre: S/N, marca, modelo, cliente
      ?status=<val>     → filtra pelo status operacional (choice)
      ?empresa=<texto>  → filtra pelo nome_fantasia do cliente (icontains)
      ?modelo=<texto>   → filtra pelo modelo do equipamento (icontains)
      ?export=excel     → dispara o download do relatório Excel filtrado
    """
    # ── 1. Captura dos filtros ────────────────────────────────────────────────
    f_q       = request.GET.get("q",       "").strip()
    f_status  = request.GET.get("status",  "").strip()
    f_empresa = request.GET.get("empresa", "").strip()
    f_modelo  = request.GET.get("modelo",  "").strip()

    # ── 2. QuerySet base ──────────────────────────────────────────────────────
    ativos = Ativo.objects.select_related(
        "modelo_equipamento", "cliente_atual"
    ).order_by("-data_cadastro")

    # ── 3. Aplicação encadeada dos filtros ────────────────────────────────────
    if f_status:
        ativos = ativos.filter(status=f_status)

    if f_empresa:
        ativos = ativos.filter(
            cliente_atual__nome_fantasia__icontains=f_empresa
        )

    if f_modelo:
        ativos = ativos.filter(
            Q(modelo_equipamento__marca__icontains=f_modelo)
            | Q(modelo_equipamento__modelo__icontains=f_modelo)
        )

    if f_q:
        ativos = ativos.filter(
            Q(numero_serie__icontains=f_q)
            | Q(modelo_equipamento__marca__icontains=f_q)
            | Q(modelo_equipamento__modelo__icontains=f_q)
            | Q(cliente_atual__nome_fantasia__icontains=f_q)
            | Q(unidade__icontains=f_q)
            | Q(setor__icontains=f_q)
        )

    # ── 4. KPIs REATIVOS — calculados sobre o QuerySet já filtrado ────────────
    #   Usamos uma única query com aggregate para minimizar roundtrips ao banco.
    kpis = ativos.aggregate(
        total      = Count("pk"),
        kpi_estoque   = Count("pk", filter=Q(status=Ativo.Status.ESTOQUE)),
        kpi_alocado   = Count("pk", filter=Q(status=Ativo.Status.ALOCADO)),
        kpi_manutencao= Count("pk", filter=Q(status=Ativo.Status.MANUTENCAO)),
        kpi_descarte  = Count("pk", filter=Q(status=Ativo.Status.DESCARTE)),
    )

    # ── 5. Exportação Excel (?export=excel) ────────────────────────────────────
    if request.GET.get("export") == "excel":
        return _exportar_excel(ativos)

    # ── 6. Listas para os dropdowns de filtro por coluna ──────────────────────
    empresas_disponiveis = (
        Ativo.objects
        .exclude(cliente_atual__isnull=True)
        .values_list("cliente_atual__nome_fantasia", flat=True)
        .order_by("cliente_atual__nome_fantasia")
        .distinct()
    )
    modelos_disponiveis = (
        ModeloEquipamento.objects
        .values_list("marca", "modelo")
        .order_by("marca", "modelo")
        .distinct()
    )

    # ── 7. Flag: há algum filtro ativo? ───────────────────────────────────────
    filtros_ativos = any([f_q, f_status, f_empresa, f_modelo])

    context = {
        "ativos"              : ativos,
        "total"               : kpis["total"],
        "kpi_estoque"         : kpis["kpi_estoque"],
        "kpi_alocado"         : kpis["kpi_alocado"],
        "kpi_manutencao"      : kpis["kpi_manutencao"],
        "kpi_descarte"        : kpis["kpi_descarte"],
        "status_choices"      : Ativo.Status.choices,
        # Valores atuais dos filtros (para manter o estado nos inputs)
        "f_q"                 : f_q,
        "f_status"            : f_status,
        "f_empresa"           : f_empresa,
        "f_modelo"            : f_modelo,
        "filtros_ativos"      : filtros_ativos,
        # Listas para os dropdowns
        "empresas_disponiveis": empresas_disponiveis,
        "modelos_disponiveis" : modelos_disponiveis,
    }
    return render(request, "inventario/lista_ativos.html", context)


def _exportar_excel(ativos_qs):
    """
    Gera e retorna um arquivo Excel (.xlsx) com os ativos do queryset recebido.
    Usa openpyxl — instale com: pip install openpyxl
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "❌ Biblioteca openpyxl não instalada. Execute: pip install openpyxl",
            status=500,
            content_type="text/plain",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventário de Ativos"

    # Cabeçalho
    cabecalho = [
        "ID", "Número de Série", "Marca", "Modelo", "Categoria",
        "Status", "Cliente / Empresa", "Unidade", "Setor", "Data de Cadastro",
    ]
    header_fill = PatternFill(start_color="1A0036", end_color="1A0036", fill_type="solid")
    header_font = Font(bold=True, color="C084F5", size=10)

    ws.append(cabecalho)
    for col_idx, _ in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dados
    for ativo in ativos_qs:
        ws.append([
            ativo.pk,
            ativo.numero_serie,
            ativo.modelo_equipamento.marca,
            ativo.modelo_equipamento.modelo,
            ativo.modelo_equipamento.get_categoria_display(),
            ativo.get_status_display(),
            ativo.cliente_atual.nome_fantasia if ativo.cliente_atual else "",
            ativo.unidade or "",
            ativo.setor or "",
            ativo.data_cadastro.strftime("%d/%m/%Y %H:%M") if ativo.data_cadastro else "",
        ])

    # Auto-largura das colunas
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # Retorna como download
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="inventario_ativos.xlsx"'
    return response




@login_required(login_url="login")
def novo_ativo(request):
    """
    Renderiza e processa o formulário de cadastro de um novo Ativo físico.

    Em caso de sucesso, redireciona para a lista de ativos com mensagem flash.
    """
    if request.method == "POST":
        form = AtivoForm(request.POST)
        if form.is_valid():
            ativo = form.save()
            messages.success(
                request,
                f"✅ Ativo <strong>S/N {ativo.numero_serie}</strong> cadastrado com sucesso!",
            )
            return redirect("inventario:lista_ativos")
    else:
        form = AtivoForm()

    return render(request, "inventario/novo_ativo.html", {"form": form, "titulo": "Novo Ativo"})


@login_required(login_url="login")
def editar_ativo(request, pk):
    """
    Renderiza e processa o formulário de edição de um Ativo existente.
    Reutiliza o mesmo template de cadastro (novo_ativo.html).
    """
    ativo = get_object_or_404(Ativo, pk=pk)

    if request.method == "POST":
        form = AtivoForm(request.POST, instance=ativo)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                f"✅ Ativo <strong>S/N {ativo.numero_serie}</strong> atualizado com sucesso!",
            )
            return redirect("inventario:lista_ativos")
    else:
        form = AtivoForm(instance=ativo)

    return render(
        request,
        "inventario/novo_ativo.html",
        {"form": form, "titulo": f"Editar Ativo — S/N {ativo.numero_serie}", "ativo": ativo},
    )


@login_required(login_url="login")
def deletar_ativo(request, pk):
    """
    Exclui um Ativo via POST. Ações GET são ignoradas (segurança).
    """
    if request.method == "POST":
        ativo = get_object_or_404(Ativo, pk=pk)
        sn = ativo.numero_serie
        ativo.delete()
        messages.success(request, f"🗑️ Ativo S/N {sn} removido do inventário.")
    return redirect("inventario:lista_ativos")


# ─────────────────────────────────────────────────────────────────────────────
# CATÁLOGO DE MODELOS
# ─────────────────────────────────────────────────────────────────────────────


@login_required(login_url="login")
def cadastrar_modelo(request):
    """
    Renderiza e processa o formulário de cadastro de um novo modelo
    no catálogo de hardware (ModeloEquipamento).
    """
    if request.method == "POST":
        form = ModeloEquipamentoForm(request.POST)
        if form.is_valid():
            modelo = form.save()
            messages.success(
                request,
                f"✅ Modelo <strong>{modelo}</strong> adicionado ao catálogo!",
            )
            return redirect("inventario:lista_ativos")
    else:
        form = ModeloEquipamentoForm()

    return render(
        request,
        "inventario/cadastrar_modelo.html",
        {"form": form, "titulo": "Novo Modelo no Catálogo"},
    )


@login_required(login_url="login")
def lista_modelos(request):
    """
    Lista todos os modelos cadastrados no catálogo de hardware.
    """
    modelos = ModeloEquipamento.objects.all().order_by("marca", "modelo")

    # Fallback simples com contagem via prefetch
    from django.db.models import Count
    modelos = ModeloEquipamento.objects.annotate(
        qtd_ativos=Count("ativos")
    ).order_by("marca", "modelo")

    return render(
        request,
        "inventario/lista_modelos.html",
        {"modelos": modelos, "titulo": "Catálogo de Modelos"},
    )


# ─────────────────────────────────────────────────────────────────────────────
# BIPAGEM EM LOTE
# ─────────────────────────────────────────────────────────────────────────────

import json
from django.db import IntegrityError
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods


@login_required(login_url="login")
@require_http_methods(["GET", "POST"])
def bipagem_lote(request):
    """
    Tela de inventário em lote (Kaizen — eliminação de desperdício).

    GET  → renderiza a página com os selects de contexto.
    POST → endpoint AJAX que recebe {modelo_id, status, unidade, setor,
           numero_serie} e salva um Ativo, retornando JSON.

    Tratamentos de erro incluídos:
      - numero_serie duplicado  → IntegrityError → HTTP 409
      - modelo_id inválido      → HTTP 404
      - campos obrigatórios     → HTTP 400
    """
    if request.method == "GET":
        context = {
            "modelos": ModeloEquipamento.objects.order_by("marca", "modelo"),
            "status_choices": Ativo.Status.choices,
        }
        return render(request, "inventario/bipagem_lote.html", context)

    # ── POST: salvar um Ativo via AJAX ────────────────────────────────────
    try:
        payload = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"ok": False, "erro": "Payload inválido."}, status=400)

    numero_serie  = payload.get("numero_serie", "").strip()
    modelo_id     = payload.get("modelo_id")
    status        = payload.get("status", Ativo.Status.ESTOQUE)
    unidade       = payload.get("unidade", "").strip()
    setor         = payload.get("setor", "").strip()

    # Validações básicas
    if not numero_serie:
        return JsonResponse({"ok": False, "erro": "Número de série não pode ser vazio."}, status=400)
    if not modelo_id:
        return JsonResponse({"ok": False, "erro": "Selecione um modelo antes de bipar."}, status=400)

    try:
        modelo = ModeloEquipamento.objects.get(pk=modelo_id)
    except ModeloEquipamento.DoesNotExist:
        return JsonResponse({"ok": False, "erro": "Modelo não encontrado."}, status=404)

    try:
        ativo = Ativo.objects.create(
            modelo_equipamento=modelo,
            numero_serie=numero_serie,
            status=status,
            unidade=unidade or None,
            setor=setor or None,
        )
    except IntegrityError:
        return JsonResponse(
            {"ok": False, "erro": f"S/N «{numero_serie}» já está cadastrado no sistema."},
            status=409,
        )

    return JsonResponse(
        {
            "ok": True,
            "ativo_id": ativo.pk,
            "numero_serie": ativo.numero_serie,
            "modelo": str(modelo),
            "status": ativo.get_status_display(),
        },
        status=201,
    )


# ─────────────────────────────────────────────────────────────────────────────
# MOVIMENTAÇÃO TABLET E CENTRAL
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="login")
def lista_movimentacoes(request):
    """
    Central de Custódia e Logística de ativos.
    Lista o histórico de movimentações com filtros.
    """
    movimentacoes = Movimentacao.objects.select_related("cliente", "usuario_responsavel").prefetch_related("itens").order_by("-data")

    # Filtros
    f_cliente = request.GET.get("cliente", "")
    f_tipo = request.GET.get("tipo", "")
    f_data_inicio = request.GET.get("data_inicio", "")
    f_data_fim = request.GET.get("data_fim", "")

    if f_cliente:
        movimentacoes = movimentacoes.filter(cliente_id=f_cliente)
    if f_tipo:
        movimentacoes = movimentacoes.filter(tipo=f_tipo)
    if f_data_inicio:
        movimentacoes = movimentacoes.filter(data__date__gte=f_data_inicio)
    if f_data_fim:
        movimentacoes = movimentacoes.filter(data__date__lte=f_data_fim)

    # KPIs do mês (simples)
    from django.utils import timezone
    mes_atual = timezone.now().month
    ano_atual = timezone.now().year
    
    kpis = {
        "alocacoes_mes": Movimentacao.objects.filter(tipo=Movimentacao.Tipo.ALOCACAO, data__month=mes_atual, data__year=ano_atual).count(),
        "devolucoes_mes": Movimentacao.objects.filter(tipo=Movimentacao.Tipo.DEVOLUCAO, data__month=mes_atual, data__year=ano_atual).count(),
        "trocas_mes": Movimentacao.objects.filter(tipo=Movimentacao.Tipo.TROCA, data__month=mes_atual, data__year=ano_atual).count(),
    }

    clientes = Cliente.objects.all().order_by("nome_fantasia")
    tipos = Movimentacao.Tipo.choices

    return render(request, "inventario/movimentacao_list.html", {
        "movimentacoes": movimentacoes,
        "clientes": clientes,
        "tipos_movimentacao": tipos,
        "kpis": kpis,
        "f_cliente": f_cliente,
        "f_tipo": f_tipo,
        "f_data_inicio": f_data_inicio,
        "f_data_fim": f_data_fim,
    })

@login_required(login_url="login")
def tela_movimentacao(request):
    """
    Renderiza a tela otimizada para tablet (bipagem de movimentações).
    """
    clientes = Cliente.objects.all().order_by("nome_fantasia")
    tipos = Movimentacao.Tipo.choices
    return render(request, "inventario/movimentacao_tablet.html", {
        "clientes": clientes,
        "tipos_movimentacao": tipos
    })

@login_required(login_url="login")
@require_http_methods(["GET"])
def validar_ativo_bipagem(request):
    """
    Valida via AJAX se um ativo pode ser bipado para o tipo de movimentação selecionado.
    """
    numero_serie = request.GET.get("numero_serie", "").strip()
    tipo = request.GET.get("tipo", "")

    if not numero_serie:
        return JsonResponse({"ok": False, "erro": "Número de série vazio."})

    try:
        ativo = Ativo.objects.get(numero_serie=numero_serie)
    except Ativo.DoesNotExist:
        return JsonResponse({"ok": False, "erro": f"S/N {numero_serie} não encontrado."})

    # Regras de Negócio Básicas
    if tipo == Movimentacao.Tipo.ALOCACAO and ativo.status != Ativo.Status.ESTOQUE:
        return JsonResponse({"ok": False, "erro": f"Ativo {numero_serie} não está em Estoque (Status: {ativo.get_status_display()})."})
    
    if tipo == Movimentacao.Tipo.TESTE and ativo.status != Ativo.Status.ESTOQUE:
        return JsonResponse({"ok": False, "erro": f"Ativo {numero_serie} não está em Estoque (Status: {ativo.get_status_display()})."})
    
    if tipo in [Movimentacao.Tipo.DEVOLUCAO, Movimentacao.Tipo.TROCA] and ativo.status not in [Ativo.Status.ALOCADO, Ativo.Status.TESTE]:
        return JsonResponse({"ok": False, "erro": f"Ativo {numero_serie} não está Alocado ou em Teste (Status: {ativo.get_status_display()})."})

    return JsonResponse({
        "ok": True,
        "ativo_id": ativo.pk,
        "modelo": f"{ativo.modelo_equipamento.marca} {ativo.modelo_equipamento.modelo}",
        "status_atual": ativo.get_status_display(),
        "cliente_atual": ativo.cliente_atual.nome_fantasia if ativo.cliente_atual else None
    })

@login_required(login_url="login")
@require_http_methods(["GET"])
def buscar_acessorios(request):
    """
    Autocomplete AJAX: retorna acessórios (controla_serial=False) com estoque > 0.
    """
    q = request.GET.get("q", "").strip()
    if len(q) < 2:
        return JsonResponse({"resultados": []})

    acessorios = Ativo.objects.filter(
        controla_serial=False,
        quantidade_estoque__gt=0,
    ).filter(
        Q(modelo_equipamento__marca__icontains=q) |
        Q(modelo_equipamento__modelo__icontains=q) |
        Q(modelo_equipamento__categoria__icontains=q)
    ).select_related("modelo_equipamento")[:15]

    resultados = [{
        "id": a.pk,
        "nome": f"{a.modelo_equipamento.marca} {a.modelo_equipamento.modelo}",
        "categoria": a.modelo_equipamento.get_categoria_display(),
        "estoque_disponivel": a.quantidade_estoque,
    } for a in acessorios]

    return JsonResponse({"resultados": resultados})


@login_required(login_url="login")
@require_http_methods(["POST"])
def processar_lote_movimentacao(request):
    """
    Processa a movimentação de um lote de ativos.
    Suporta itens com serial (bipados) e acessórios de lote (sem serial, com quantidade).
    """
    try:
        payload = json.loads(request.body)
        cliente_id = payload.get("cliente_id")
        tipo = payload.get("tipo")
        observacoes = payload.get("observacoes", "")
        itens = payload.get("itens", [])
        acessorios = payload.get("acessorios", [])
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "erro": "Payload JSON inválido."}, status=400)

    if not cliente_id or not tipo or (not itens and not acessorios):
        return JsonResponse({"ok": False, "erro": "Faltam dados obrigatórios (itens ou acessórios)."}, status=400)

    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return JsonResponse({"ok": False, "erro": "Cliente não encontrado."}, status=404)

    ativos_atualizados = []

    try:
        with transaction.atomic():
            # Cria a Movimentação Principal
            movimentacao = Movimentacao.objects.create(
                tipo=tipo,
                cliente=cliente,
                usuario_responsavel=request.user,
                observacoes=observacoes
            )

            # ── Itens COM serial (bipados) ─────────────────────────────────
            for item_data in itens:
                sn_principal = item_data.get("numero_serie")
                ativo_principal = Ativo.objects.select_for_update().get(numero_serie=sn_principal)

                # Troca (Swap) lida com dois ativos
                if tipo == Movimentacao.Tipo.TROCA:
                    sn_substituto = item_data.get("numero_serie_substituto")
                    ativo_substituto = Ativo.objects.select_for_update().get(numero_serie=sn_substituto)
                    
                    # O ativo defeituoso volta pro estoque como TRIAGEM
                    ativo_principal.status = Ativo.Status.TRIAGEM
                    ativo_principal.cliente_atual = None
                    ativo_principal.save()

                    # O ativo novo vai para o cliente como ALOCADO
                    ativo_substituto.status = Ativo.Status.ALOCADO
                    ativo_substituto.cliente_atual = cliente
                    ativo_substituto.save()

                    MovimentacaoItem.objects.create(
                        movimentacao=movimentacao,
                        ativo=ativo_principal,
                        ativo_substituto=ativo_substituto
                    )
                    ativos_atualizados.extend([ativo_principal, ativo_substituto])

                else:
                    if tipo == Movimentacao.Tipo.ALOCACAO:
                        ativo_principal.status = Ativo.Status.ALOCADO
                        ativo_principal.cliente_atual = cliente
                    elif tipo == Movimentacao.Tipo.TESTE:
                        ativo_principal.status = Ativo.Status.TESTE
                        ativo_principal.cliente_atual = cliente
                    elif tipo == Movimentacao.Tipo.DEVOLUCAO:
                        ativo_principal.status = Ativo.Status.TRIAGEM
                        ativo_principal.cliente_atual = None
                    # Se for VISITA_TECNICA, não altera o status
                    
                    ativo_principal.save()
                    MovimentacaoItem.objects.create(
                        movimentacao=movimentacao,
                        ativo=ativo_principal
                    )
                    ativos_atualizados.append(ativo_principal)

            # ── Acessórios SEM serial (lote com quantidade) ────────────────
            for acessorio_data in acessorios:
                acessorio_id = acessorio_data.get("ativo_id")
                qtd = int(acessorio_data.get("quantidade", 0))

                if qtd <= 0:
                    continue

                acessorio = Ativo.objects.select_for_update().get(pk=acessorio_id, controla_serial=False)

                if tipo in [Movimentacao.Tipo.ALOCACAO, Movimentacao.Tipo.TESTE, Movimentacao.Tipo.VISITA]:
                    # Saída de estoque: decrementar
                    if acessorio.quantidade_estoque < qtd:
                        raise ValueError(
                            f"Estoque insuficiente para '{acessorio.modelo_equipamento}'. "
                            f"Disponível: {acessorio.quantidade_estoque}, Solicitado: {qtd}."
                        )
                    acessorio.quantidade_estoque -= qtd
                elif tipo == Movimentacao.Tipo.DEVOLUCAO:
                    # Entrada no estoque: incrementar
                    acessorio.quantidade_estoque += qtd

                acessorio.save()

                MovimentacaoItem.objects.create(
                    movimentacao=movimentacao,
                    ativo=acessorio,
                    quantidade=qtd
                )
            
            # Gera o PDF usando xhtml2pdf
            import os
            from django.conf import settings
            context = {
                "movimentacao": movimentacao,
                "itens": movimentacao.itens.all(),
                "logo_path": os.path.join(settings.BASE_DIR, 'imagens', 'aletechlgo.png').replace('\\', '/')
            }
            pdf_bytes = render_to_pdf("inventario/cautela_pdf.html", context)
            if pdf_bytes:
                filename = f"cautela_{movimentacao.id}_{movimentacao.tipo}.pdf"
                movimentacao.cautela_pdf.save(filename, ContentFile(pdf_bytes))

    except Ativo.DoesNotExist:
        return JsonResponse({"ok": False, "erro": "Um dos ativos não foi encontrado no banco."}, status=400)
    except ValueError as e:
        return JsonResponse({"ok": False, "erro": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "erro": str(e)}, status=500)

    pdf_url = movimentacao.cautela_pdf.url if movimentacao.cautela_pdf else None

    return JsonResponse({
        "ok": True,
        "mensagem": "Movimentação registrada com sucesso!",
        "pdf_url": pdf_url
    })


@login_required(login_url="login")
def gerar_pdf_cautela(request, movimentacao_id):
    """
    Gera dinamicamente o PDF do termo de cautela da movimentação solicitada.
    """
    import os
    from django.conf import settings
    from django.http import HttpResponse

    movimentacao = get_object_or_404(Movimentacao, pk=movimentacao_id)
    
    context = {
        "movimentacao": movimentacao,
        "itens": movimentacao.itens.all(),
        # Caminho absoluto do logo PNG para o xhtml2pdf (via link_callback)
        "logo_path": os.path.join(settings.BASE_DIR, 'imagens', 'aletechlgo.png').replace('\\', '/')
    }
    
    pdf_bytes = render_to_pdf("inventario/cautela_pdf.html", context)
    
    if pdf_bytes:
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        filename = f"cautela_{movimentacao.id}_{movimentacao.tipo}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
        
    return HttpResponse("Erro ao gerar o PDF.", status=500)

